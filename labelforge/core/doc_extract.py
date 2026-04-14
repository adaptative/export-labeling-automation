"""Extract text from uploaded documents for AI classification.

Supports PDF (via PyMuPDF/fitz) and XLSX (via openpyxl).
Falls back to raw UTF-8 decode for other file types.
"""
from __future__ import annotations

import logging
from typing import Optional

logger = logging.getLogger(__name__)


def extract_text(data: bytes, filename: str, max_chars: int = 3000) -> str:
    """Extract readable text from a document's raw bytes.

    Args:
        data: Raw file bytes from blob store.
        filename: Original filename (used to pick the right parser).
        max_chars: Truncate output to this length.

    Returns:
        Extracted text, or empty string on failure.
    """
    lower = filename.lower()
    try:
        if lower.endswith(".pdf"):
            return _extract_pdf(data, max_chars)
        elif lower.endswith((".xlsx", ".xls")):
            return _extract_xlsx(data, max_chars)
        else:
            return data.decode("utf-8", errors="replace")[:max_chars]
    except Exception as exc:
        logger.warning("Text extraction failed for %s: %s", filename, exc)
        return ""


def _extract_pdf(data: bytes, max_chars: int) -> str:
    """Extract text from PDF using PyMuPDF."""
    import fitz  # PyMuPDF

    text_parts: list[str] = []
    total = 0
    with fitz.open(stream=data, filetype="pdf") as doc:
        for page in doc:
            page_text = page.get_text()
            text_parts.append(page_text)
            total += len(page_text)
            if total >= max_chars:
                break

    return "\n".join(text_parts)[:max_chars]


def _extract_xlsx(data: bytes, max_chars: int) -> str:
    """Extract text from XLSX using openpyxl."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    text_parts: list[str] = []
    total = 0

    for sheet in wb.worksheets[:3]:  # First 3 sheets max
        text_parts.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(max_row=50, values_only=True):  # First 50 rows
            cells = [str(c).replace("\n", " ").replace("\r", " ").strip() if c is not None else "" for c in row]
            line = "\t".join(cells)
            text_parts.append(line)
            total += len(line)
            if total >= max_chars:
                break
        if total >= max_chars:
            break

    wb.close()
    return "\n".join(text_parts)[:max_chars]
