"""Tests for labelforge.core.doc_extract — text extraction from documents."""
from __future__ import annotations

import pytest

from labelforge.core.doc_extract import extract_text


class TestExtractTextPlainFiles:
    """Plain text / fallback extraction."""

    def test_plain_text_file(self):
        content = b"Hello world\nLine two\nLine three"
        result = extract_text(content, "notes.txt")
        assert "Hello world" in result
        assert "Line two" in result

    def test_plain_text_truncated_by_max_chars(self):
        content = b"A" * 500
        result = extract_text(content, "big.txt", max_chars=100)
        assert len(result) <= 100

    def test_binary_file_decoded_with_replace(self):
        content = b"\xff\xfe\x00Invalid UTF8"
        result = extract_text(content, "data.bin")
        # Should not raise — uses errors="replace"
        assert isinstance(result, str)

    def test_empty_file(self):
        result = extract_text(b"", "empty.txt")
        assert result == ""

    def test_csv_treated_as_plain_text(self):
        content = b"col1,col2\nval1,val2\n"
        result = extract_text(content, "data.csv")
        assert "col1,col2" in result


class TestExtractTextPDF:
    """PDF extraction via PyMuPDF."""

    def test_valid_pdf(self):
        """A minimal valid PDF with embedded text should return that text."""
        # Minimal valid PDF with text
        import fitz
        doc = fitz.open()
        page = doc.new_page(width=200, height=200)
        page.insert_text((50, 100), "Hello PDF World")
        pdf_bytes = doc.tobytes()
        doc.close()

        result = extract_text(pdf_bytes, "test.pdf")
        assert "Hello PDF World" in result

    def test_corrupt_pdf_returns_empty(self):
        """Corrupt PDF data should return empty string, not raise."""
        result = extract_text(b"not a real pdf", "broken.pdf")
        assert result == ""

    def test_pdf_max_chars_truncation(self):
        import fitz
        doc = fitz.open()
        page = doc.new_page()
        page.insert_text((50, 100), "X" * 500)
        pdf_bytes = doc.tobytes()
        doc.close()

        result = extract_text(pdf_bytes, "big.pdf", max_chars=50)
        assert len(result) <= 50


class TestExtractTextXLSX:
    """XLSX extraction via openpyxl."""

    def test_valid_xlsx(self):
        import io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Item"
        ws["B1"] = "Qty"
        ws["A2"] = "Widget"
        ws["B2"] = 100
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        result = extract_text(xlsx_bytes, "order.xlsx")
        assert "Item" in result
        assert "Widget" in result

    def test_xlsx_max_chars(self):
        import io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in range(1, 100):
            ws.cell(row=row, column=1, value="A" * 50)
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        result = extract_text(xlsx_bytes, "big.xlsx", max_chars=200)
        assert len(result) <= 200

    def test_xls_extension_also_works(self):
        """Even with .xls extension, openpyxl path should be tried."""
        # openpyxl can't read real .xls, so this should fail gracefully
        result = extract_text(b"not real xls", "old.xls")
        assert result == ""
